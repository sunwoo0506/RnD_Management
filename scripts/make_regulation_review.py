# 규정 DB 패키지 폴더의 JSON 산출물을 검토용 엑셀(Review.xlsx)로 묶는다.
# 사용법: python scripts/make_regulation_review.py <패키지 폴더>
#
# 시트 구성은 사업 종류와 무관하게 고정한다 (docs/gwayeon_guideline_extraction_framework/04_mvp_output_spec.md §4.3):
#   Summary / BudgetTree / BudgetGuides / AllowedItems / LimitRules / RuleReview
# 사업마다 시트명이 다르면 검토 절차를 사업 수만큼 만들어야 하므로 이름과 순서를 바꾸지 않는다.
#
# 입력 파일명은 MVP 규격(expense_categories, source_text, regulation_rules)을 우선하고,
# 없으면 이전 이름(legal_budget_tree, regulation_articles, approval/evidence/applicability)으로 읽는다.
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

pkg = Path(sys.argv[1])
out = pkg / "Review.xlsx"


def load(*names):
    """이름 후보를 순서대로 찾아 첫 번째로 존재하는 파일을 읽는다 (MVP 이름 우선)."""
    for name in names:
        path = pkg / name
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    return None


manifest = load("manifest.json") or {}
# 문서 메타(문서명·고시번호·시행일)는 MVP 규격상 manifest에 들어간다.
# 이전 패키지는 document.json으로 분리돼 있어 없으면 그쪽에서 읽는다.
document = load("document.json") or {}


def doc_meta(*keys, default=""):
    for key in keys:
        for source in (manifest, document):
            value = source.get(key)
            if value:
                return value
    return default
tree = load("expense_categories.json", "legal_budget_tree.json") or []
guides = load("budget_screen_guides.json") or []
items = load("expense_allowed_items.json") or []
limits = load("expense_limit_rules.json") or []
articles = load("source_text.json", "regulation_articles.json") or []
# 판정 규칙: MVP는 regulation_rules 한 파일. 이전 패키지는 승인·증빙·적용조건이 흩어져 있다.
rules = load("regulation_rules.json") or []
approvals = load("approval_rules.json") or []
evidence = load("evidence_rules.json") or []
applicability = load("expense_applicability_rules.json") or []

wb = Workbook()
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")


def sheet(title, headers, rows, widths):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
    for row in rows:
        ws.append(["" if v is None else (", ".join(map(str, v)) if isinstance(v, list) else v) for v in row])
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.font = Font(size=10)
    ws.freeze_panes = "A2"
    return ws


# ---- 1. Summary — 규정 기본사항 ----
# 시행일이 조문마다 다른 경우(부칙 경과조치)를 반드시 드러낸다.
def effective_date_notes():
    """조문마다 시행일이 다른 경우를 "제5조·제10조의2: 2026-06-11 시행" 형태로 만든다.
    문서 메타의 special_effective_dates가 있으면 그것을 쓰고(부칙에서 직접 뽑은 값이라 정확하다),
    없으면 조문별 effective_from이 기본 시행일과 다른 것을 모은다."""
    special = doc_meta("special_effective_dates", default=[])
    if special:
        return [("·".join(s.get("articles", [])), f"{s.get('effective_from')} 시행"
                 + (f" ({s['reason']})" if s.get("reason") else "")) for s in special]
    base = doc_meta("effective_from")
    by_date = {}
    for a in articles:
        eff = a.get("effective_from")
        ref = a.get("source_article") or a.get("location")
        if eff and ref and eff != base:
            by_date.setdefault(eff, []).append(ref)
    return [(f"{'·'.join(refs[:8])}{' 외' if len(refs) > 8 else ''}", f"{date} 시행") for date, refs in sorted(by_date.items())]


ws = wb.active
ws.title = "Summary"
counts = manifest.get("counts", {})
meta = manifest.get("pack_meta", {})
ws.append([f"과제온 규정 DB 검토본 — {manifest.get('document_version', pkg.name)}"])
ws["A1"].font = Font(bold=True, size=13)
ws.append([])
summary_rows = [
    ("문서", meta.get("guideline") or doc_meta("title", "document_title", "document_version")),
    ("고시·공고 번호", doc_meta("notice_number", "version_label")),
    ("발행기관", meta.get("agency") or doc_meta("issuer", "agency")),
    ("문서 유형", doc_meta("doc_type", "document_type")),
    ("개정 구분", doc_meta("revision_type")),
    ("시행", doc_meta("effective_from", "promulgated_at")),
]
for ref, note in effective_date_notes():
    summary_rows.append((ref, note))
summary_rows += [
    ("베이스 지침", manifest.get("base_document_version", "(없음)")),
    ("생성일", doc_meta("generated_at")),
    ("원본 파일", " / ".join(manifest.get("source_files", [])) or doc_meta("source_filename")),
    ("비고", manifest.get("notes", "")),
    ("", ""),
    ("비목 노드", counts.get("legal_categories", len(tree))),
    ("허용상한 규칙", counts.get("limit_rules", len(limits))),
    ("사용 가능 항목", counts.get("allowed_items", len(items))),
    ("화면 가이드", counts.get("budget_guides", len(guides))),
    ("판정 규칙", len(rules) + len(approvals) + len(evidence) + len(applicability)),
    ("조문 원문", len(articles)),
]
for key, value in summary_rows:
    ws.append([key, "" if value is None else value])
    ws[ws.max_row][0].font = Font(bold=True, size=10)
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 90
for row in ws.iter_rows(min_row=3):
    row[1].alignment = WRAP

# ---- 2. BudgetTree — 법정 비목의 계층구조 ----
sheet(
    "BudgetTree",
    ["비목 코드", "비목명", "상위 비목", "구분", "레벨", "최하위", "하위 수", "사용 항목 수", "정렬"],
    [(n.get("category_code") or n.get("code"), n.get("category_name") or n.get("name"), n.get("parent_code"),
      n.get("cost_class"), n.get("level"), str(n.get("is_leaf_category", "")), n.get("child_category_count"),
      n.get("allowed_item_count"), n.get("display_order")) for n in tree],
    [26, 22, 16, 10, 6, 8, 8, 12, 6],
)

# ---- 3. BudgetGuides — 화면에 표시할 비목별 사용 가능 항목·허용상한 (검토용 한글 헤더) ----
sheet(
    "BudgetGuides",
    ["비목 코드", "화면명", "사용 요약", "허용 상한", "상세 기준", "근거", "시행일"],
    [(g.get("category_code"), g.get("display_name"), g.get("usage_summary"), g.get("limit_text"),
      g.get("limit_detail_text"), g.get("source_articles"), g.get("effective_from")) for g in guides],
    [22, 16, 48, 28, 54, 26, 11],
)

# ---- 4. AllowedItems — 비목 아래 실제 사용할 수 있는 항목 ----
sheet(
    "AllowedItems",
    ["항목 코드", "비목 코드", "사용 가능 항목", "설명", "적용 기관", "가용 상태", "조건", "제한", "사전승인", "증빙", "근거"],
    [(i.get("item_code"), i.get("category_code"), i.get("item_name"), i.get("description"),
      i.get("institution_scope"), i.get("availability_status"), i.get("condition_summary"),
      i.get("restriction_summary"), str(i.get("requires_approval", False)), i.get("evidence_summary"),
      i.get("source_article")) for i in items],
    [24, 24, 26, 46, 14, 14, 34, 34, 9, 28, 24],
)

# ---- 5. LimitRules — 금액·비율·계산식·승인·인정 규칙 ----
# MVP limit_type 7종(04_mvp_output_spec.md §4.2)으로 정규화해 함께 보여준다.
MVP_LIMIT_TYPE = {
    "NO_FIXED_CAP": "NONE", "NONE": "NONE", "PLAN_BASED": "NONE", "ELIGIBILITY": "NONE",
    "FIXED_AMOUNT": "FIXED_AMOUNT", "PER_TRANSACTION": "FIXED_AMOUNT", "PER_PERSON": "FIXED_AMOUNT", "PER_PERIOD": "FIXED_AMOUNT",
    "PERCENT": "PERCENT", "CHANGE_RATE": "PERCENT",
    "FORMULA": "FORMULA", "ANNUAL_AVERAGE": "ANNUAL_AVERAGE",
    "APPROVAL_THRESHOLD": "APPROVAL_THRESHOLD", "PROCEDURE_THRESHOLD": "APPROVAL_THRESHOLD", "APPROVAL": "APPROVAL_THRESHOLD",
    "RECOGNITION_LIMIT": "RECOGNITION_LIMIT",
}


def mvp_limit_type(rule):
    raw = rule.get("limit_type")
    # 초과 시 기관 인정이 필요한 규칙은 승인 임계값과 구분한다
    if rule.get("over_limit_action") == "RECOGNITION_REQUIRED" and raw not in ("PERCENT", "FIXED_AMOUNT", "FORMULA"):
        return "RECOGNITION_LIMIT"
    return MVP_LIMIT_TYPE.get(raw, raw)


sheet(
    "LimitRules",
    ["규칙 코드", "비목 코드", "규칙명", "원본 유형", "MVP 유형", "값", "단위", "산정 기준", "계산식", "화면 문구", "초과 처리", "적용 기관", "근거", "원문 인용"],
    [(r.get("limit_code"), r.get("category_code"), r.get("limit_name"), r.get("limit_type"), mvp_limit_type(r),
      r.get("limit_value"), r.get("limit_unit"), r.get("basis_ko") or r.get("basis_code"), r.get("formula_expression"),
      r.get("ui_summary"), r.get("over_limit_action"), r.get("institution_scope"), r.get("source_article"),
      r.get("source_quote")) for r in limits],
    [24, 22, 24, 18, 18, 10, 9, 24, 22, 44, 20, 14, 24, 54],
)

# ---- 6. RuleReview — 자동판정 규칙과 상세 판정 규칙 ----
# 상한(LimitRules)은 자동 판정 가능 여부를, 승인·증빙·적용조건은 상세 판정 규칙으로 한 시트에 모은다.
# 화면 문구와 원문 인용을 함께 실어, 이 시트만 보고도 "사용자에게 이렇게 보이는데 원문이 이렇다"를
# 대조할 수 있게 한다 (검토자가 다른 시트를 오갈 필요가 없도록).


def rule_message(r):
    """사용자 화면에 나가는 문구. 패키지에 따라 ui_summary 또는 result.message에 들어 있다."""
    result = r.get("result")
    if isinstance(result, dict) and result.get("message"):
        return result["message"]
    return r.get("ui_summary") or r.get("rule_name")


def norm_ref(text):
    return re.sub(r"[\s·.,()]", "", str(text or ""))


# 조문 번호 → 원문. source_text.json을 색인해 두고, 인용이 없는 규칙의 근거를 채운다.
ARTICLE_BY_REF = {}
for _a in articles:
    _ref = _a.get("source_article") or _a.get("location")
    _text = _a.get("original_text") or _a.get("text")
    if _ref and _text:
        ARTICLE_BY_REF.setdefault(norm_ref(_ref), _text)

ARTICLE_EXCERPT = 400


def article_text_for(source_article):
    """근거에 적힌 조문 번호로 원문을 찾는다. "제27조제3항·제73조제1항제7호"처럼 여러 조문이면 각각."""
    if not source_article:
        return None
    found = []
    for m in re.finditer(r"제\s*\d+(?:의\s*\d+)?\s*조", str(source_article)):
        text = ARTICLE_BY_REF.get(norm_ref(m.group(0)))
        if text and text not in found:
            found.append(text)
    if not found:
        text = ARTICLE_BY_REF.get(norm_ref(source_article))
        if text:
            found.append(text)
    if not found:
        return None
    joined = "\n".join(found)
    return joined if len(joined) <= ARTICLE_EXCERPT else joined[:ARTICLE_EXCERPT] + " …(조문 전문은 source_text 참조)"


def rule_quote(r):
    """근거 원문. source_quote(인용) 우선, 없으면 source_text(조문 본문),
    그것도 없으면 근거 조문 번호로 원문에서 찾아 채운다 (NRD 상한 규칙은 인용 필드가 아예 없다)."""
    return r.get("source_quote") or r.get("source_text") or article_text_for(r.get("source_article"))


def review_status(r):
    """원문 대조 상태. 규칙이 직접 인용을 갖고 있는지, 조문만 연결된 것인지 구분한다."""
    if r.get("source_quote") or r.get("source_text"):
        return "SOURCE_VERIFIED"
    if article_text_for(r.get("source_article")):
        return "ARTICLE_LINKED"
    return "NEEDS_QUOTE"


review_rows = []
for r in limits:
    computable = r.get("limit_type") in ("PERCENT", "FORMULA") and r.get("limit_value") is not None
    review_rows.append((
        r.get("limit_code"), r.get("limit_name"), r.get("category_code"), "상한",
        "AUTO_CAP" if computable else "MANUAL_REVIEW",
        "BLOCKING" if r.get("over_limit_action") == "NOT_ALLOWED" else "WARNING",
        r.get("ui_summary"), rule_quote(r),
        r.get("source_article"), r.get("effective_from"), review_status(r),
    ))
for r in rules:
    review_rows.append((
        r.get("rule_code"), r.get("rule_name"), r.get("expense_category_code") or r.get("category_code"),
        r.get("rule_type") or "판정", r.get("automation_level") or "MANUAL_REVIEW",
        (r.get("result") or {}).get("status") if isinstance(r.get("result"), dict) else r.get("result"),
        rule_message(r), rule_quote(r),
        r.get("source_article"), r.get("effective_from"), review_status(r),
    ))
for r in approvals:
    review_rows.append((r.get("approval_code"), r.get("rule_name"), r.get("category_code"), "사전승인·인정",
                        "MANUAL_REVIEW", r.get("result_status"), rule_message(r), rule_quote(r),
                        r.get("source_article"), r.get("effective_from"), review_status(r)))
for r in evidence:
    review_rows.append((r.get("evidence_code"), r.get("rule_name"), r.get("category_code"), "증빙",
                        "SEMI_AUTO", ", ".join(r.get("required_documents", [])),
                        rule_message(r), rule_quote(r),
                        r.get("source_article"), r.get("effective_from"), review_status(r)))
for r in applicability:
    review_rows.append((r.get("applicability_code") or r.get("rule_code"), r.get("condition_summary") or r.get("condition"),
                        r.get("category_code"), "적용 조건", "MANUAL_REVIEW", r.get("result"),
                        r.get("condition_summary") or r.get("condition"), rule_quote(r),
                        r.get("source_article"), r.get("effective_from"), review_status(r)))
sheet(
    "RuleReview",
    ["규칙 코드", "규칙명", "비목", "규칙 구분", "자동화", "판정 결과", "화면 문구", "원문 인용", "근거", "시행일", "검토 상태"],
    review_rows,
    [22, 34, 22, 14, 15, 22, 48, 60, 24, 11, 15],
)

wb.save(out)
# 윈도우 콘솔(cp949)에서도 깨지지 않게 출력한다
def say(text):
    try:
        print(text)
    except UnicodeEncodeError:
        sys.stdout.buffer.write(text.encode("utf-8", "replace") + b"\n")


say(f"{out} - 시트 {len(wb.sheetnames)}개: {', '.join(wb.sheetnames)}")
say(f"  비목 {len(tree)} / 가이드 {len(guides)} / 항목 {len(items)} / 상한 {len(limits)} / 판정 {len(review_rows)} / 조문 {len(articles)}")
